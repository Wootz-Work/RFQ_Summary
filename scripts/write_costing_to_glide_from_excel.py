from __future__ import annotations

import argparse
import sys
import time
from pathlib import Path

import httpx
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from rfq_summary.config import load_settings


WORKBOOK_DEFAULT = ROOT / "fc0465.All RFQs - Strike.xlsx"
RFQ_SHEET = "fc0465.All RFQs - Strike"


def cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def header_map(ws) -> dict[str, int]:
    return {cell_text(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}


def require_col(headers: dict[str, int], name: str) -> int:
    if name not in headers:
        raise KeyError(f"Missing column {name!r}. Found: {list(headers)}")
    return headers[name]


def require_row_id_col(headers: dict[str, int]) -> int:
    for name, col in headers.items():
        if "row id" in name.lower():
            return col
    raise KeyError(f"Missing row id column. Found: {list(headers)}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Write Excel costing order-of-magnitude values to Glide ALL RFQ without touching other columns."
    )
    parser.add_argument("--workbook", default=str(WORKBOOK_DEFAULT))
    parser.add_argument("--dry-run", action="store_true", help="Show planned Glide updates without writing.")
    parser.add_argument("--limit", type=int, default=0, help="Write at most N rows.")
    return parser.parse_args()


def write_costing_value(settings, row_id: str, estimate: str) -> None:
    if not settings.glide_api_key or not settings.glide_app_id:
        raise RuntimeError("Missing GLIDE_API_KEY/GLIDE_APP_ID.")
    if not settings.glide_all_rfq_table:
        raise RuntimeError("Missing GLIDE_ALL_RFQ_TABLE.")

    estimate_col = (settings.glide_col_all_rfq_costing_order_of_magnitude or "").strip()
    if not estimate_col:
        raise RuntimeError("Missing GLIDE_COL_ALL_RFQ_COSTING_ORDER_OF_MAGNITUDE.")

    payload = {
        "appID": settings.glide_app_id,
        "mutations": [
            {
                "kind": "set-columns-in-row",
                "tableName": settings.glide_all_rfq_table,
                "rowID": row_id.strip(),
                "columnValues": {
                    estimate_col: estimate,
                },
            }
        ],
    }
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {settings.glide_api_key}",
    }
    with httpx.Client(timeout=60) as client:
        response = client.post("https://api.glideapp.io/api/function/mutateTables", headers=headers, json=payload)
        response.raise_for_status()


def main() -> int:
    args = parse_args()
    workbook_path = Path(args.workbook).expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[RFQ_SHEET]
    headers = header_map(ws)
    row_id_col = require_row_id_col(headers)
    title_col = require_col(headers, "Title")
    estimate_col = require_col(headers, "Costing order of magnitude")

    planned: list[tuple[int, str, str, str]] = []
    for row in range(2, ws.max_row + 1):
        row_id = cell_text(ws.cell(row, row_id_col).value)
        estimate = cell_text(ws.cell(row, estimate_col).value)
        title = cell_text(ws.cell(row, title_col).value)
        if not row_id or not estimate:
            continue
        planned.append((row, row_id, title, estimate))
        if args.limit and len(planned) >= args.limit:
            break

    print(f"Workbook: {workbook_path.name}")
    print(f"Rows with non-empty costing values: {len(planned)}")
    print("Will update only the Glide costing column; no other columns are included in the mutation.")
    for row, row_id, title, estimate in planned:
        print(f"  row={row} row_id={row_id} estimate={estimate!r} title={title!r}")

    if args.dry_run or not planned:
        return 0

    settings = load_settings()
    failures: list[tuple[int, str, str]] = []
    for index, (row, row_id, title, estimate) in enumerate(planned, 1):
        t0 = time.perf_counter()
        try:
            write_costing_value(settings, row_id, estimate)
            elapsed = int((time.perf_counter() - t0) * 1000)
            print(f"[{index}/{len(planned)}] row={row} row_id={row_id} estimate={estimate!r} written in {elapsed}ms")
        except Exception as exc:
            failures.append((row, row_id, f"{type(exc).__name__}: {exc}"))
            print(f"[{index}/{len(planned)}] row={row} row_id={row_id} failed: {type(exc).__name__}: {exc}")

    if failures:
        print("Failures:")
        for row, row_id, err in failures:
            print(f"  row={row} row_id={row_id}: {err}")
        return 1

    print("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
