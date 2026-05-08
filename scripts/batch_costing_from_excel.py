from __future__ import annotations

import argparse
import re
import shutil
import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from rfq_summary.attachments import analyze_attachments
from rfq_summary.config import load_settings
from rfq_summary.llm import generate_text, load_prompt_file
from rfq_summary.schema import QueryPayload
from rfq_summary.task import _build_query_triage_prompt, _join_attachment_text_any, _unwrap_tagged_output


WORKBOOK_DEFAULT = ROOT / "fc0465.All RFQs - Strike.xlsx"
RFQ_SHEET = "fc0465.All RFQs - Strike"
PRODUCT_SHEET = "ALL Product"
SKIP_DWG_CUSTOMERS = {"pragati", "prakriti", "unnati"}
URL_RE = re.compile(r"https?://[^\s,<>\"]+")
DRIVE_ID_RE = re.compile(r"^[A-Za-z0-9_-]{25,44}$")


def cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def header_map(ws) -> dict[str, int]:
    return {cell_text(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}


def require_col(headers: dict[str, int], name: str) -> int:
    if name not in headers:
        raise KeyError(f"Missing column {name!r}. Found: {list(headers)}")
    return headers[name]


def require_row_id_col(headers: dict[str, int]) -> int:
    for name, col in headers.items():
        if "row id" in name.lower():
            return col
    raise KeyError(f"Missing row id column. Found: {list(headers)}")


def extract_links(*cells) -> list[str]:
    links: list[str] = []
    for cell in cells:
        if cell is None:
            continue
        if getattr(cell, "hyperlink", None) and cell.hyperlink.target:
            links.append(cell.hyperlink.target)
        text = cell_text(cell.value)
        links.extend(URL_RE.findall(text))
        if DRIVE_ID_RE.fullmatch(text):
            links.append(text)
    return dedupe(links)


def dedupe(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        value = cell_text(value).rstrip(").,]")
        if value and value not in seen:
            seen.add(value)
            out.append(value)
    return out


def build_product_index(product_ws, product_cols: dict[str, int]) -> dict[str, list[int]]:
    rfq_col = require_col(product_cols, "rfq id")
    by_rfq: dict[str, list[int]] = defaultdict(list)
    for row in range(2, product_ws.max_row + 1):
        rfq_id = cell_text(product_ws.cell(row, rfq_col).value)
        if rfq_id:
            by_rfq[rfq_id].append(row)
    return by_rfq


def build_body(rfq_ws, product_ws, rfq_cols: dict[str, int], product_cols: dict[str, int], row: int, product_rows: list[int]) -> str:
    lines = [
        f"Title: {cell_text(rfq_ws.cell(row, require_col(rfq_cols, 'Title')).value)}",
        f"Customer: {cell_text(rfq_ws.cell(row, require_col(rfq_cols, 'Customer name')).value)}",
        f"Industry: {cell_text(rfq_ws.cell(row, require_col(rfq_cols, 'Industry')).value)}",
        f"Geography: {cell_text(rfq_ws.cell(row, require_col(rfq_cols, 'Geography')).value)}",
        f"Standard: {cell_text(rfq_ws.cell(row, require_col(rfq_cols, 'Standard')).value)}",
        f"Deadline: {cell_text(rfq_ws.cell(row, require_col(rfq_cols, 'Deadline')).value)}",
        f"Current status: {cell_text(rfq_ws.cell(row, require_col(rfq_cols, 'Current status')).value)}",
    ]
    if product_rows:
        lines.append("Products:")
        for product_row in product_rows:
            name = cell_text(product_ws.cell(product_row, require_col(product_cols, "Name")).value)
            qty = cell_text(product_ws.cell(product_row, require_col(product_cols, "Qty")).value)
            details = cell_text(product_ws.cell(product_row, require_col(product_cols, "Details")).value)
            lines.append(f"- Name: {name}; Qty: {qty}; Details: {details}")
    return "\n".join(line for line in lines if line.strip())


def build_attachment_urls(
    rfq_ws,
    product_ws,
    rfq_cols: dict[str, int],
    product_cols: dict[str, int],
    row: int,
    product_rows: list[int],
) -> list[str]:
    customer = cell_text(rfq_ws.cell(row, require_col(rfq_cols, "Customer name")).value).lower()
    skip_dwg = customer in SKIP_DWG_CUSTOMERS

    urls: list[str] = []
    urls.extend(extract_links(rfq_ws.cell(row, require_col(rfq_cols, "Quotation folder link"))))

    for product_row in product_rows:
        if not skip_dwg:
            urls.extend(extract_links(product_ws.cell(product_row, require_col(product_cols, "Dwg link"))))
        urls.extend(extract_links(product_ws.cell(product_row, require_col(product_cols, "Rep URL"))))
        urls.extend(extract_links(product_ws.cell(product_row, require_col(product_cols, "Addl. files Internal (non active projects) copy"))))

    return dedupe(urls)


def generate_costing(settings, payload: QueryPayload) -> tuple[str, str]:
    attachment_findings = analyze_attachments(settings, payload.all_attachment_urls())
    extracted_text = _join_attachment_text_any("", attachment_findings)
    prompt_template = load_prompt_file(settings.prompt_query_costing_file)
    user_prompt = _build_query_triage_prompt(prompt_template, payload, extracted_text)
    raw = generate_text(
        settings,
        system_prompt="You must follow the user instructions exactly.",
        user_prompt=user_prompt,
    )
    return _unwrap_tagged_output(raw, "estimate"), raw


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fill Costing order of magnitude in the RFQ Excel workbook.")
    parser.add_argument("--workbook", default=str(WORKBOOK_DEFAULT))
    parser.add_argument("--force", action="store_true", help="Regenerate rows even when column already has a value.")
    parser.add_argument("--dry-run", action="store_true", help="Print planned rows and attachment counts without calling the LLM or saving.")
    parser.add_argument("--limit", type=int, default=0, help="Process at most N rows.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook_path = Path(args.workbook).expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)

    wb = load_workbook(workbook_path)
    rfq_ws = wb[RFQ_SHEET]
    product_ws = wb[PRODUCT_SHEET]
    rfq_cols = header_map(rfq_ws)
    product_cols = header_map(product_ws)
    row_id_col = require_row_id_col(rfq_cols)
    output_col = require_col(rfq_cols, "Costing order of magnitude")
    product_index = build_product_index(product_ws, product_cols)

    planned: list[tuple[int, str, QueryPayload]] = []
    for row in range(2, rfq_ws.max_row + 1):
        existing = cell_text(rfq_ws.cell(row, output_col).value)
        if existing and not args.force:
            continue
        rfq_id = cell_text(rfq_ws.cell(row, row_id_col).value)
        if not rfq_id:
            continue
        product_rows = product_index.get(rfq_id, [])
        payload = QueryPayload(
            row_id=rfq_id,
            subject=cell_text(rfq_ws.cell(row, require_col(rfq_cols, "Title")).value),
            from_="",
            from_name=cell_text(rfq_ws.cell(row, require_col(rfq_cols, "Customer name")).value),
            body=build_body(rfq_ws, product_ws, rfq_cols, product_cols, row, product_rows),
            received_at="",
            attachment_urls=build_attachment_urls(rfq_ws, product_ws, rfq_cols, product_cols, row, product_rows),
            attached_media=[],
        )
        planned.append((row, rfq_id, payload))
        if args.limit and len(planned) >= args.limit:
            break

    print(f"Workbook: {workbook_path.name}")
    print(f"Rows to process: {len(planned)}")
    for row, rfq_id, payload in planned:
        print(f"  row={row} rfq_id={rfq_id} attachments={len(payload.attachment_urls)} subject={payload.subject!r}")

    if args.dry_run or not planned:
        return 0

    backup = workbook_path.with_name(f"{workbook_path.stem}.backup-{datetime.now().strftime('%Y%m%d-%H%M%S')}{workbook_path.suffix}")
    shutil.copy2(workbook_path, backup)
    print(f"Backup: {backup.name}")

    settings = load_settings()
    failures: list[tuple[int, str, str]] = []
    for index, (row, rfq_id, payload) in enumerate(planned, 1):
        print(f"[{index}/{len(planned)}] row={row} rfq_id={rfq_id} starting...")
        t0 = time.perf_counter()
        try:
            estimate, raw = generate_costing(settings, payload)
            rfq_ws.cell(row, output_col).value = estimate
            wb.save(workbook_path)
            elapsed = int((time.perf_counter() - t0) * 1000)
            print(f"[{index}/{len(planned)}] row={row} estimate={estimate!r} saved in {elapsed}ms")
            if not estimate:
                print(f"[{index}/{len(planned)}] row={row} raw output was empty estimate: {raw!r}")
        except Exception as exc:
            elapsed = int((time.perf_counter() - t0) * 1000)
            failures.append((row, rfq_id, f"{type(exc).__name__}: {exc}"))
            print(f"[{index}/{len(planned)}] row={row} failed in {elapsed}ms: {type(exc).__name__}: {exc}")

    if failures:
        print("Failures:")
        for row, rfq_id, err in failures:
            print(f"  row={row} rfq_id={rfq_id}: {err}")
        return 1

    print("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
